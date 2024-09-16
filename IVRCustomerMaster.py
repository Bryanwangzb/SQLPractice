"""
Title: Data Process Script
Description: This script processes raw data and creates csv files.
Author: WangZhibin
Date Created: 2024-06-02
Last Modified: 2024-06-21
Version: 1.0
"""

import csv
import datetime
import json
import logging
import os
import sys
import time
from logging import getLogger, config

import mysql.connector
import numpy as np
import pandas as pd
import xlwings as xw
from mysql.connector import Error
import configparser
import win32com.client as win32
import openpyxl
from openpyxl.styles import Font

import calendar

with open('./log_config.json', 'r') as f:
    log_conf = json.load(f)

config.dictConfig(log_conf)

# Log Setting
logger = getLogger(__name__)

# settings initial file
inifile = configparser.ConfigParser()
inifile.read('settings.ini')

db_section = 'AWS_SLAVE'

# get db and user info from inifile.
db_host_name = inifile.get(db_section, 'db_host_name')
db_user_name = inifile.get(db_section, 'db_user_name')

# [LOCAL]Database Schema
schema_dev = inifile.get(db_section, 'schema_dev')
schema_prod = inifile.get(db_section, 'schema_prod')

# DBアカウント情報
pwd = inifile.get(db_section, 'pwd')
db = schema_prod

# Script directory
script_directory = os.path.dirname(os.path.abspath(sys.argv[0]))

# 日付情報
running_date = datetime.date.today()
# 　日付時刻情報
timestr = time.strftime("%Y%m%d%H%M%S")

# get the number of days in given month
previous_month_days = calendar.monthrange(2024,5)[1]
current_month_days = calendar.monthrange(2024,6)[1]

# get the numver of weeks in given month
previous_month_weeks = len(calendar.monthcalendar(2024,5))
current_month_weeks = len(calendar.monthcalendar(2024,6))


# データベースに接続
def create_db_connection(host_name, user_name, user_password, db_name):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password,
            database=db_name
        )
        logger.info("MySQL Database connection successful")
    except Error as err:
        logger.error(f"Error: '{err}'")

    return connection


# データの読み取り
def read_query(connection, query):
    cursor = connection.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        return result
    except Error as err:
        logger.error(f"Error: '{err}'")


# create connect to database
conn = create_db_connection(db_host_name, db_user_name, pwd, db)

#########
# Query
#########
q_company_amount = f'''
SELECT
    count(*) 
FROM
    companies
where
    date_sub(created_at,interval 9 hour) < '2024/07/01'
'''

q_ivr_customer = f'''WITH user_registration AS ( 
    SELECT
        com.id AS company_id
        , com.name AS company_name
        , count(cu.id) AS user_amount 
    FROM
        companies AS com 
    LEFT OUTER JOIN company_users AS cu 
        ON com.id = cu.company_id 
    where
        date_sub(cu.created_at,interval 9 hour)<'2024/07/01'
    GROUP BY
        com.id
) 
, new_user_registration AS ( 
    SELECT
        company_id
        , count(*) AS user_new 
    FROM
        company_users 
    WHERE
        DATE_SUB(created_at,interval 9 hour) >= '2024/07/01' - INTERVAL 1 MONTH 
        AND DATE_SUB(created_at,interval 9 hour) < '2024/07/01' 
    GROUP BY
        company_id
) 
, daily_login_amount AS ( 
    WITH daiy_login_temp AS ( 
        SELECT
            cu.company_id AS company_id
            , count(DISTINCT  DATE_SUB(cl.created_at,interval 9 hour)) AS login_count 
        FROM
            company_logs AS cl 
            LEFT OUTER JOIN company_users AS cu 
                ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
        WHERE
            cl.url = 'api/company/auth/login' 
            AND cu.email NOT LIKE '%@brownreverse%' 
            AND  DATE_SUB(cl.created_at,interval 9 hour) >= ('2024/07/01' - INTERVAL 1 MONTH) 
            AND  DATE_SUB(cl.created_at,interval 9 hour) < '2024/07/01'  
        GROUP BY
            cu.id
            ,  DATE_SUB(cl.created_at,interval 9 hour)
    ) 
    SELECT
        company_id
        , sum(login_count) / {current_month_days} AS daily_login_amount
    FROM
        daiy_login_temp
    GROUP BY company_id
) 
, daily_login_previous_month AS ( 
    WITH daily_login_temp AS ( 
        SELECT
            cu.company_id AS company_id
            ,  DATE_SUB(cl.created_at,interval 9 hour) AS login_date
            , COUNT(DISTINCT  DATE_SUB(cl.created_at,interval 9 hour)) AS login_count 
        FROM
            company_logs AS cl 
            LEFT OUTER JOIN company_users AS cu 
                ON JSON_UNQUOTE(JSON_EXTRACT(cl.request_parameters, '$.email')) = cu.email 
        WHERE
            cl.url = 'api/company/auth/login' 
            AND cu.email NOT LIKE '%@brownreverse%' 
        GROUP BY
              cu.id
            ,  DATE_SUB(cl.created_at,interval 9 hour)
    ) 
    , login_counts AS ( 
        SELECT
            company_id
            , SUM( 
                CASE 
                    WHEN login_date >= ('2024/06/01' - INTERVAL 1 MONTH) 
                    AND login_date < '2024/06/01' 
                        THEN login_count 
                    ELSE 0 
                    END
            ) AS previous_month_login_count
            , SUM( 
                CASE 
                    WHEN login_date >= ('2024/07/01' - INTERVAL 1 MONTH) 
                    AND login_date < '2024/07/01' 
                        THEN login_count 
                    ELSE 0 
                    END
            ) AS current_month_login_count 
        FROM
            daily_login_temp 
        GROUP BY
            company_id
    ) 
    SELECT
        company_id,
        CASE 
            WHEN current_month_login_count = 0 
                THEN NULL 
            ELSE (current_month_login_count / {current_month_days}) / (previous_month_login_count / {previous_month_days})
            END AS daily_login_count_ratio 
    FROM
        login_counts
) 
, weekly_login_amount AS ( 
    WITH week_login_temp AS ( 
        SELECT
            cu.company_id AS company_id
            ,  yearweek(DATE_SUB(cl.created_at,interval 9 hour)) AS created_yearweek
            , count(DISTINCT  yearweek(DATE_SUB(cl.created_at,interval 9 hour))) AS login_count 
        FROM
            company_logs AS cl 
            LEFT OUTER JOIN company_users AS cu 
                ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
        WHERE
            cl.url = 'api/company/auth/login' 
            AND cu.email NOT LIKE '%@brownreverse%' 
            AND  DATE_SUB(cl.created_at,interval 9 hour) >= ('2024/07/01' - INTERVAL 1 MONTH) 
            AND  DATE_SUB(cl.created_at,interval 9 hour) < '2024/07/01' 
        GROUP BY
            cu.id
            ,  yearweek(DATE_SUB(cl.created_at,interval 9 hour))
    ) 
    SELECT
        company_id
        , sum(login_count) / {current_month_weeks} AS weekly_login_amount
    FROM
        week_login_temp
    GROUP BY company_id
) 
, weekly_login_previous_month AS ( 
    WITH year_week_current_login AS ( 
        WITH week_login_temp AS ( 
            SELECT
                cu.company_id AS company_id
                , cu.id
                ,  yearweek(DATE_SUB(cl.created_at,interval 9 hour)) AS created_yearweek
                , count(DISTINCT  yearweek(DATE_SUB(cl.created_at,interval 9 hour))) AS login_count 
            FROM
                company_logs AS cl 
                LEFT OUTER JOIN company_users AS cu 
                    ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
            WHERE
                cl.url = 'api/company/auth/login' 
                AND cu.email NOT LIKE '%@brownreverse%' 
                AND  DATE_SUB(cl.created_at,interval 9 hour) >= '2024/07/01' - INTERVAL 1 MONTH 
                AND  DATE_SUB(cl.created_at,interval 9 hour) < '2024/07/01' 
            GROUP BY
                cu.id
                ,  yearweek(DATE_SUB(cl.created_at,interval 9 hour))
        ) 
        SELECT
            company_id
            , sum(login_count) AS temp_result_current
            , sum(login_count) / {current_month_weeks} AS RESULT
        FROM
            week_login_temp
        GROUP BY company_id
    ) 
    , year_week_previous_login AS ( 
        WITH week_login_temp AS ( 
            SELECT
                cu.company_id AS company_id
                , cu.id
                ,  yearweek(DATE_SUB(cl.created_at,interval 9 hour)) AS created_yearweek
                , count(DISTINCT  yearweek(DATE_SUB(cl.created_at,interval 9 hour))) AS login_count 
            FROM
                company_logs AS cl 
                LEFT OUTER JOIN company_users AS cu 
                    ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
            WHERE
                cl.url = 'api/company/auth/login' 
                AND cu.email NOT LIKE '%@brownreverse%' 
                AND  DATE_SUB(cl.created_at,interval 9 hour) >= ('2024/06/01' - INTERVAL 1 MONTH) 
                AND  DATE_SUB(cl.created_at,interval 9 hour) < '2024/06/01' 
            GROUP BY
                cu.id
                ,  yearweek(DATE_SUB(cl.created_at,interval 9 hour))
        ) 
        SELECT
            company_id
            , sum(login_count) AS temp_result_previous
            , sum(login_count) / {current_month_weeks} AS RESULT 
        FROM
            week_login_temp
        GROUP BY company_id
    ) 
    SELECT
        ywc.company_id 
        , ywc.result / ywp.result AS weekly_login_count_ratio
    FROM
        year_week_current_login AS ywc 
        LEFT OUTER JOIN year_week_previous_login AS ywp 
            ON ywc.company_id = ywp.company_id
        order by ywc.company_id
)
 
, monthly_login_amount AS ( 
    WITH month_login_temp AS ( 
        SELECT
            cu.company_id AS company_id
            ,  yearweek(DATE_SUB(cl.created_at,interval 9 hour)) AS created_yearweek
            , count(DISTINCT MONTH (DATE_SUB(cl.created_at,interval 9 hour))) AS login_count 
        FROM
            company_logs AS cl 
            LEFT OUTER JOIN company_users AS cu 
                ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
        WHERE
            cl.url = 'api/company/auth/login' 
            AND cu.email NOT LIKE '%@brownreverse%' 
            AND  DATE_SUB(cl.created_at,interval 9 hour) >= ('2024/07/01' - INTERVAL 1 MONTH) 
            AND  DATE_SUB(cl.created_at,interval 9 hour) < '2024/07/01' 
        GROUP BY
            cu.id
            , MONTH (DATE_SUB(cl.created_at,interval 9 hour))
    ) 
    SELECT
        company_id
        , sum(login_count) AS monthly_login_amount
    FROM
        month_login_temp
    GROUP BY company_id
) 
, monthly_login_previous_month AS ( 
    WITH year_month_current_login AS ( 
        WITH month_login_temp AS ( 
            SELECT
                cu.company_id AS company_id
                ,  yearweek(DATE_SUB(cl.created_at,interval 9 hour)) AS created_yearweek
                , count(DISTINCT MONTH (DATE_SUB(cl.created_at,interval 9 hour))) AS login_count 
            FROM
                company_logs AS cl 
                LEFT OUTER JOIN company_users AS cu 
                    ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
            WHERE
                cl.url = 'api/company/auth/login' 
                AND cu.email NOT LIKE '%@brownreverse%' 
                AND  DATE_SUB(cl.created_at,interval 9 hour) >= ('2024/07/01' - INTERVAL 1 MONTH) 
                AND  DATE_SUB(cl.created_at,interval 9 hour) < '2024/07/01' 
            GROUP BY
                cu.id
                , MONTH (DATE_SUB(cl.created_at,interval 9 hour))
        ) 
        SELECT
            company_id
            , sum(login_count) AS login_count 
        FROM
            month_login_temp
        GROUP BY 
            company_id
    ) 
    , year_month_previous_login AS ( 
        WITH month_login_temp AS ( 
            SELECT
                cu.company_id AS company_id
                ,  yearweek(DATE_SUB(cl.created_at,interval 9 hour)) AS created_yearweek
                , count(DISTINCT MONTH (DATE_SUB(cl.created_at,interval 9 hour))) AS login_count 
            FROM
                company_logs AS cl 
                LEFT OUTER JOIN company_users AS cu 
                    ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
            WHERE
                cl.url = 'api/company/auth/login' 
                AND cu.email NOT LIKE '%@brownreverse%' 
                AND  DATE_SUB(cl.created_at,interval 9 hour) >= ('2024/06/01' - INTERVAL 1 MONTH) 
                AND  DATE_SUB(cl.created_at,interval 9 hour) < '2024/06/01' 
            GROUP BY
                cu.id
                , MONTH (DATE_SUB(cl.created_at,interval 9 hour))
        ) 
        SELECT
            company_id
            , sum(login_count) AS login_count 
        FROM
            month_login_temp
        GROUP BY company_id
    ) 
    SELECT
        ymc.company_id
        , ymc.login_count / ymp.login_count AS monthly_login_previous_month_ratio
    FROM
        year_month_current_login AS ymc 
        LEFT OUTER JOIN year_month_previous_login AS ymp 
            ON ymc.company_id = ymp.company_id
) 
, company_area_info AS ( 
    SELECT
        com.id AS company_id
        , com.name AS company_name
        , pa.id AS area_id
        , pa.name AS area_name 
    FROM
        companies AS com 
        LEFT OUTER JOIN plants AS pla 
            ON com.id = pla.company_id 
        LEFT OUTER JOIN plant_areas AS pa 
            ON pla.id = pa.plant_id
) 
, marker_amount AS ( 
    SELECT
        cai.company_id
        , count(*) AS marker_amount
    FROM
        markers AS mk 
        LEFT OUTER JOIN company_area_info AS cai 
            ON mk.plant_area_id = cai.area_id 
    WHERE
         date_sub(mk.created_at, interval 9 hour) < '2024/07/01' and cai.company_id is not null
    GROUP by cai.company_id
) 
, new_marker_amount AS ( 
    SELECT
        cai.company_id
        , count(*) AS new_marker_amount
    FROM
        markers AS mk 
        LEFT OUTER JOIN company_area_info AS cai 
            ON mk.plant_area_id = cai.area_id 
    WHERE
        date_sub(mk.created_at, interval 9 hour) < '2024/07/01' 
        AND date_sub(mk.created_at, interval 9 hour) >= DATE_SUB('2024/07/01', INTERVAL 1 MONTH) 
    GROUP BY
        cai.company_id
) 
, machine_amount AS ( 
    SELECT
        cai.company_id
        , count(*) AS machine_amount
    FROM
        assets AS ast 
        LEFT OUTER JOIN company_area_info AS cai 
            ON ast.plant_area_id = cai.area_id 
    WHERE
        date_sub(ast.created_at, interval 9 hour) < '2024/07/01' 
        AND cai.company_id IS NOT NULL 
    GROUP BY
        cai.company_id
) 
, new_machine_amount AS ( 
    SELECT
        cai.company_id
        , count(*) AS new_machine_amount
    FROM
        assets AS ast 
        LEFT OUTER JOIN company_area_info AS cai 
            ON ast.plant_area_id = cai.area_id 
    WHERE
        date_sub(ast.created_at, interval 9 hour) < '2024/07/01' 
        AND date_sub(ast.created_at, interval 9 hour) >= DATE_SUB('2024/07/01', INTERVAL 1 MONTH) 
        AND cai.company_id IS NOT NULL 
    GROUP BY
        cai.company_id
) 
, measurement_amount AS ( 
    SELECT
        cai.company_id
        , count(*) AS measurement_amount
    FROM
        measure_lengths AS ms 
        LEFT OUTER JOIN company_area_info AS cai 
            ON ms.plant_area_id = cai.area_id 
    WHERE
        date_sub(ms.created_at, interval 9 hour) < '2024/07/01' 
        AND cai.company_id IS NOT NULL 
    GROUP BY
        cai.company_id
) 
, new_measurement_amount AS ( 
    SELECT
        cai.company_id
        , count(*) AS new_measurement_amount
    FROM
        measure_lengths AS ms 
        LEFT OUTER JOIN company_area_info AS cai 
            ON ms.plant_area_id = cai.area_id 
    WHERE
        date_sub(ms.created_at, interval 9 hour) < '2024/07/01' 
        AND date_sub(ms.created_at, interval 9 hour) >= DATE_SUB('2024/07/01', INTERVAL 1 MONTH) 
        AND cai.company_id IS NOT NULL 
    GROUP BY
        cai.company_id
) 
, simulation_amount AS ( 
    SELECT
        cai.company_id
        , count(*) AS simulation_amount
    FROM
        plant_area_objects AS pao 
        LEFT OUTER JOIN company_area_info AS cai 
            ON pao.plant_area_id = cai.area_id 
    WHERE
        date_sub(pao.created_at, interval 9 hour) < '2024/07/01' 
        AND cai.company_id IS NOT NULL 
    GROUP BY
        cai.company_id
) 
, new_simulation_amount AS ( 
    SELECT
        cai.company_id
        , count(*) AS new_simulation_amount
    FROM
        plant_area_objects AS pao 
        LEFT OUTER JOIN company_area_info AS cai 
            ON pao.plant_area_id = cai.area_id 
    WHERE
        date_sub(pao.created_at, interval 9 hour) < '2024/07/01' 
        AND date_sub(pao.created_at, interval 9 hour) >= DATE_SUB('2024/07/01', INTERVAL 1 MONTH) 
        AND cai.company_id IS NOT NULL 
    GROUP BY
        cai.company_id
) 
, pipenavi_amount AS ( 
    SELECT
        cai.company_id
        , count(*) AS pipenavi_amount
    FROM
        pipe_groups AS pg 
        LEFT OUTER JOIN company_area_info AS cai 
            ON pg.plant_area_id = cai.area_id 
    WHERE
        date_sub(pg.created_at, interval 9 hour) < '2024/07/01' 
        AND cai.company_id IS NOT NULL 
    GROUP BY
        cai.company_id
) 
, new_pipenavi_amount AS ( 
    SELECT
        cai.company_id
        , count(*) AS new_pipenavi_amount
    FROM
        pipe_groups AS pg 
        LEFT OUTER JOIN company_area_info AS cai 
            ON pg.plant_area_id = cai.area_id 
    WHERE
        date_sub(pg.created_at, interval 9 hour) < '2024/07/01' 
        AND date_sub(pg.created_at, interval 9 hour) >= DATE_SUB('2024/07/01', INTERVAL 1 MONTH) 
        AND cai.company_id IS NOT NULL 
    GROUP BY
        cai.company_id
) 
SELECT 
    ur.company_name as '会社'
    , ur.user_amount
    , nur.user_new 
    , dla.daily_login_amount
    , concat(round(dlp.daily_login_count_ratio*100,0),'%')
    , wla.weekly_login_amount
    , concat(round(wlp.weekly_login_count_ratio*100,0),'%')
    , mla.monthly_login_amount 
    , concat(round(mlp.monthly_login_previous_month_ratio*100,0),'%')
    , ma.marker_amount
   , nma.new_marker_amount
   , mca.machine_amount
   , nmca.new_machine_amount
   , msa.measurement_amount
   , nmsa.new_measurement_amount
   , sma.simulation_amount
   , nsma.new_simulation_amount
   , pa.pipenavi_amount
   , npa.new_pipenavi_amount
   , ""
   , ""
   , ""
   , ""
   , ""
FROM
    user_registration AS ur 
    LEFT OUTER JOIN new_user_registration AS nur 
        ON ur.company_id = nur.company_id
    LEFT OUTER JOIN daily_login_amount AS dla
        ON ur.company_id = dla.company_id
    LEFT OUTER JOIN daily_login_previous_month AS dlp
        ON ur.company_id = dlp.company_id
    LEFT OUTER JOIN weekly_login_amount AS wla
        ON ur.company_id = wla.company_id
    LEFT OUTER JOIN weekly_login_previous_month AS wlp
        ON ur.company_id = wlp.company_id
    LEFT OUTER JOIN monthly_login_amount AS mla
        ON ur.company_id = mla.company_id
    LEFT OUTER JOIN monthly_login_previous_month AS mlp
        ON ur.company_id = mlp.company_id
    LEFT OUTER JOIN marker_amount AS ma
        ON ur.company_id = ma.company_id
    LEFT OUTER JOIN new_marker_amount AS nma
        ON ur.company_id = nma.company_id
   LEFT OUTER JOIN machine_amount AS mca
       ON ur.company_id = mca.company_id
   LEFT OUTER JOIN new_machine_amount AS nmca
       ON ur.company_id = nmca.company_id
   LEFT OUTER JOIN measurement_amount AS msa
       ON ur.company_id = msa.company_id
   LEFT OUTER JOIN new_measurement_amount AS nmsa
       ON ur.company_id = nmsa.company_id
   LEFT OUTER JOIN simulation_amount AS sma
       ON ur.company_id = sma.company_id
   LEFT OUTER JOIN new_simulation_amount AS nsma
       ON ur.company_id = nsma.company_id
   LEFT OUTER JOIN pipenavi_amount AS pa
       ON ur.company_id = pa.company_id
   LEFT OUTER JOIN new_pipenavi_amount AS npa
       ON ur.company_id = npa.company_id
'''

# ivr_customer = read_query(conn,q_ivr_customer)

company_amount = read_query(conn, q_company_amount)[0][0]

result_df = pd.read_sql_query(q_ivr_customer, conn).fillna(0)
result_df.set_index('会社', inplace=True)

result_df = result_df.stack().reset_index(level=1, drop=True).to_frame('2024年6月')
d = ["ユーザー登録数（合計）", "新規登録ユーザー数", "日間平均ログイン人数", "日間平均ログイン人数（前月比）", "週間平均ログイン"
        , "週間平均ログイン（前月比）", "月間ログイン人数", "月間ログイン人数（前月比）", "マーカー数（合計）",
     "新規マーカー数", "機番登録数（合計）", "新規機番登録数", "測長数（合計）", "新規測長数",
     "空間シミュレーション数（合計）", "新規空間シミュレーション数", "配管登録数（合計）", "新規配管登録数", "アプリDL数",
     "全体面積", "撮影面積", "撮影面積割合", "滞在時間"] * 90
result_df.insert(0, "項目", d, allow_duplicates=False)
# result_df.drop('company_id',axis=1)

# output result xlsx file.
result_df.to_excel('June_IVRCustomer.xlsx')

# excel = win32.gencache.EnsureDispatch('Excel.Application')
# wb = excel.Workbooks.Open('C:\\Users\WorkAccount\PycharmProjects\pythonProject\\venv\IVRCustomer.xlsx')
# ws = wb.Worksheets('Sheet1')

# cell_value = ws.Range('A1').Value
# print(cell_value)

workbook = openpyxl.load_workbook('./June_IVRCustomer.xlsx')
worksheet = workbook.active

# font
font = Font(name='Calibri', size=11, bold=False)

# Fill header with LightStellBlue
for cell in worksheet[1]:
    cell.fill = openpyxl.styles.PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid')
    cell.font = font

# headers = [cell.value for cell in worksheet[1]]
# for r_idx,row in enumerate(worksheet.iter_rows(min_row=2,values_only=True
#                                               ),start=2):
#     for c_idx,value in enumerate(row,start=2):
#         if headers[c_idx] == '項目' and '前月比' in str(value):
#             cell = worksheet.cell(row=r_idx,column=c_idx)
#             cell.font = Font(color="FF0000")
for row in worksheet.iter_rows(values_only=False):
    for cell in row:
        cell.font = font

item_column = worksheet['B']
for cell in item_column:
    if cell.row != 1 and '前月比' in str(cell.value) and '%' in str(
            worksheet.cell(row=cell.row, column=cell.column + 1).value):
        worksheet.cell(row=cell.row, column=cell.column + 1).font = Font(color="FF0000")

workbook.save('June_IVRCustomer.xlsx')
workbook.close()

# excel = win32.gencache.EnsureDispatch('Excel.Application')
# wb = excel.Workbooks.Open('C:\\Users\WorkAccount\PycharmProjects\pythonProject\\venv\\IVRCustomer.xlsx')
# ws = wb.Worksheets('Sheet1')
#


# cell_value = ws.Range('A1').Value
# print(cell_value)
