"""
Title: Data Process Script
Description: This script processes raw data and creates csv files.
Author: WangZhibin
Date Created: 2024-06-02
Last Modified: 2024-06-21
Version: 1.0
"""

import csv
import datetime
import json
import logging
import os
import sys
import time
from logging import getLogger, config

import mysql.connector
import numpy as np
import pandas as pd
import xlwings as xw
from mysql.connector import Error
import configparser
import win32com.client as win32
import openpyxl
from openpyxl.styles import Font

import calendar

with open('./log_config.json', 'r') as f:
    log_conf = json.load(f)

config.dictConfig(log_conf)

# Log Setting
logger = getLogger(__name__)

# settings initial file
inifile = configparser.ConfigParser()
inifile.read('settings.ini')

db_section = 'AWS_SLAVE'

# get db and user info from inifile.
db_host_name = inifile.get(db_section, 'db_host_name')
db_user_name = inifile.get(db_section, 'db_user_name')

# [LOCAL]Database Schema
schema_dev = inifile.get(db_section, 'schema_dev')
schema_prod = inifile.get(db_section, 'schema_prod')

# DBアカウント情報
pwd = inifile.get(db_section, 'pwd')
db = schema_prod

# Script directory
script_directory = os.path.dirname(os.path.abspath(sys.argv[0]))

# 日付情報
running_date = datetime.date.today()
# 　日付時刻情報
timestr = time.strftime("%Y%m%d%H%M%S")




# データベースに接続
def create_db_connection(host_name, user_name, user_password, db_name):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            passwd=user_password,
            database=db_name
        )
        logger.info("MySQL Database connection successful")
    except Error as err:
        logger.error(f"Error: '{err}'")

    return connection


# データの読み取り
def read_query(connection, query):
    cursor = connection.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        return result
    except Error as err:
        logger.error(f"Error: '{err}'")

# generate latest 1 year's year-month pair list.
def generate_year_month_pairs():
    # Get the current date
    # now = datetime.now()
    now = datetime.date.today()

    # Create a list to store (year, month) tuples
    year_month_pairs = []

    # Loop through the past 14 months including the current month
    for i in range(14):
        # Append the (year, month) tuple to the list
        year_month_pairs.append((now.year, now.month))

        # Move to the previous month
        if now.month == 1:
            now = now.replace(year=now.year - 1, month=12, day=1)
        else:
            now = now.replace(month=now.month - 1, day=1)

    # Reverse the list to get it in chronological order
    return year_month_pairs



year_month_pairs = generate_year_month_pairs()
# create connect to database
conn = create_db_connection(db_host_name, db_user_name, pwd, db)
#
# previous_month_days = calendar.monthrange(year_month_pairs[11][0], year_month_pairs[11][1])[1]
# current_month_days = calendar.monthrange(year_month_pairs[10][0], year_month_pairs[10][1])[1]
#
# # get the numver of weeks in given month
# previous_month_weeks = len(calendar.monthcalendar(year_month_pairs[11][0], year_month_pairs[11][1]))
# current_month_weeks = len(calendar.monthcalendar(year_month_pairs[10][0], year_month_pairs[10][1]))
#
# # set the month
# previous_month = str(year_month_pairs[11][0]) + "/" + str(year_month_pairs[11][1]) + "/01"
# current_month = str(year_month_pairs[10][0]) + "/" + str(year_month_pairs[10][1]) + "/01"
# year_month = str(year_month_pairs[10][0]) + "年" + str(year_month_pairs[10][1]) + "月"

# create analysis result dataframe by given current monty and previous month.
def get_analysis_result(_previous_month_days,_current_month_days,_previous_month_weeks,_current_month_weeks,_previous_month,_current_month,_year_month):
    previous_month_days = _previous_month_days
    current_month_days = _current_month_days
    previous_month_weeks = _previous_month_weeks
    current_month_weeks = _current_month_weeks
    previous_month = _previous_month
    current_month = _current_month
    year_month = _year_month
    # get the number of days in given month
    q_company_amount = f'''
    SELECT
        count(*) 
    FROM
        companies
    where
        date(created_at) < '{current_month}'
    '''

    q_ivr_customer = f'''WITH user_registration AS ( 
        SELECT
            com.id AS company_id
            , com.name AS company_name
            , count(cu.id) AS user_amount 
        FROM
            companies AS com 
        LEFT OUTER JOIN company_users AS cu 
            ON com.id = cu.company_id 
        where
            date(cu.created_at)<'{current_month}'
        GROUP BY
            com.id
    ) 
    , new_user_registration AS ( 
        SELECT
            company_id
            , count(*) AS user_new 
        FROM
            company_users 
        WHERE
            DATE (created_at) >= '{current_month}' - INTERVAL 1 MONTH 
            AND DATE (created_at) < '{current_month}' 
        GROUP BY
            company_id
    ) 
    , daily_login_amount AS ( 
        WITH daiy_login_temp AS ( 
            SELECT
                cu.company_id AS company_id
                , count(DISTINCT DATE (cl.created_at)) AS login_count 
            FROM
                company_logs AS cl 
                LEFT OUTER JOIN company_users AS cu 
                    ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
            WHERE
                cl.url = 'api/company/auth/login' 
                AND cu.email NOT LIKE '%@brownreverse%' 
                AND DATE (cl.created_at) >= ('{current_month}' - INTERVAL 1 MONTH) 
                AND DATE (cl.created_at) < '{current_month}'  
            GROUP BY
                cu.id
                , DATE (cl.created_at)
        ) 
        SELECT
            company_id
            , sum(login_count) / {current_month_days} AS daily_login_amount
        FROM
            daiy_login_temp
        GROUP BY company_id
    ) 
    , daily_login_previous_month AS ( 
        WITH daily_login_temp AS ( 
            SELECT
                cu.company_id AS company_id
                , DATE (cl.created_at) AS login_date
                , COUNT(DISTINCT DATE (cl.created_at)) AS login_count 
            FROM
                company_logs AS cl 
                LEFT OUTER JOIN company_users AS cu 
                    ON JSON_UNQUOTE(JSON_EXTRACT(cl.request_parameters, '$.email')) = cu.email 
            WHERE
                cl.url = 'api/company/auth/login' 
                AND cu.email NOT LIKE '%@brownreverse%' 
            GROUP BY
                  cu.id
                , DATE (cl.created_at)
        ) 
        , login_counts AS ( 
            SELECT
                company_id
                , SUM( 
                    CASE 
                        WHEN login_date >= ('{previous_month}' - INTERVAL 1 MONTH) 
                        AND login_date < '{previous_month}' 
                            THEN login_count 
                        ELSE 0 
                        END
                ) AS previous_month_login_count
                , SUM( 
                    CASE 
                        WHEN login_date >= ('{current_month}' - INTERVAL 1 MONTH) 
                        AND login_date < '{current_month}' 
                            THEN login_count 
                        ELSE 0 
                        END
                ) AS current_month_login_count 
            FROM
                daily_login_temp 
            GROUP BY
                company_id
        ) 
        SELECT
            company_id,
            CASE 
                WHEN current_month_login_count = 0 
                    THEN NULL 
                ELSE (current_month_login_count / {current_month_days}) / (previous_month_login_count / {previous_month_days})
                END AS daily_login_count_ratio 
        FROM
            login_counts
    ) 
    , weekly_login_amount AS ( 
        WITH week_login_temp AS ( 
            SELECT
                cu.company_id AS company_id
                , yearweek(cl.created_at) AS created_yearweek
                , count(DISTINCT yearweek(cl.created_at)) AS login_count 
            FROM
                company_logs AS cl 
                LEFT OUTER JOIN company_users AS cu 
                    ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
            WHERE
                cl.url = 'api/company/auth/login' 
                AND cu.email NOT LIKE '%@brownreverse%' 
                AND DATE (cl.created_at) >= ('{current_month}' - INTERVAL 1 MONTH) 
                AND DATE (cl.created_at) < '{current_month}' 
            GROUP BY
                cu.id
                , yearweek(cl.created_at)
        ) 
        SELECT
            company_id
            , sum(login_count) / {current_month_weeks} AS weekly_login_amount
        FROM
            week_login_temp
        GROUP BY company_id
    ) 
    , weekly_login_previous_month AS ( 
        WITH year_week_current_login AS ( 
            WITH week_login_temp AS ( 
                SELECT
                    cu.company_id AS company_id
                    , cu.id
                    , yearweek(cl.created_at) AS created_yearweek
                    , count(DISTINCT yearweek(cl.created_at)) AS login_count 
                FROM
                    company_logs AS cl 
                    LEFT OUTER JOIN company_users AS cu 
                        ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
                WHERE
                    cl.url = 'api/company/auth/login' 
                    AND cu.email NOT LIKE '%@brownreverse%' 
                    AND DATE (cl.created_at) >= '{current_month}' - INTERVAL 1 MONTH 
                    AND DATE (cl.created_at) < '{current_month}' 
                GROUP BY
                    cu.id
                    , yearweek(cl.created_at)
            ) 
            SELECT
                company_id
                , sum(login_count) AS temp_result_current
                , sum(login_count) / {current_month_weeks} AS RESULT
            FROM
                week_login_temp
            GROUP BY company_id
        ) 
        , year_week_previous_login AS ( 
            WITH week_login_temp AS ( 
                SELECT
                    cu.company_id AS company_id
                    , cu.id
                    , yearweek(cl.created_at) AS created_yearweek
                    , count(DISTINCT yearweek(cl.created_at)) AS login_count 
                FROM
                    company_logs AS cl 
                    LEFT OUTER JOIN company_users AS cu 
                        ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
                WHERE
                    cl.url = 'api/company/auth/login' 
                    AND cu.email NOT LIKE '%@brownreverse%' 
                    AND DATE (cl.created_at) >= ('{previous_month}' - INTERVAL 1 MONTH) 
                    AND DATE (cl.created_at) < '{previous_month}' 
                GROUP BY
                    cu.id
                    , yearweek(cl.created_at)
            ) 
            SELECT
                company_id
                , sum(login_count) AS temp_result_previous
                , sum(login_count) / {previous_month_weeks} AS RESULT 
            FROM
                week_login_temp
            GROUP BY company_id
        ) 
        SELECT
            ywc.company_id 
            , ywc.result / ywp.result AS weekly_login_count_ratio
        FROM
            year_week_current_login AS ywc 
            LEFT OUTER JOIN year_week_previous_login AS ywp 
                ON ywc.company_id = ywp.company_id
            order by ywc.company_id
    )

    , monthly_login_amount AS ( 
        WITH month_login_temp AS ( 
            SELECT
                cu.company_id AS company_id
                , yearweek(cl.created_at) AS created_yearweek
                , count(DISTINCT MONTH (cl.created_at)) AS login_count 
            FROM
                company_logs AS cl 
                LEFT OUTER JOIN company_users AS cu 
                    ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
            WHERE
                cl.url = 'api/company/auth/login' 
                AND cu.email NOT LIKE '%@brownreverse%' 
                AND DATE (cl.created_at) >= ('{current_month}' - INTERVAL 1 MONTH) 
                AND DATE (cl.created_at) < '{current_month}' 
            GROUP BY
                cu.id
                , MONTH (cl.created_at)
        ) 
        SELECT
            company_id
            , sum(login_count) AS monthly_login_amount
        FROM
            month_login_temp
        GROUP BY company_id
    ) 
    , monthly_login_previous_month AS ( 
        WITH year_month_current_login AS ( 
            WITH month_login_temp AS ( 
                SELECT
                    cu.company_id AS company_id
                    , yearweek(cl.created_at) AS created_yearweek
                    , count(DISTINCT MONTH (cl.created_at)) AS login_count 
                FROM
                    company_logs AS cl 
                    LEFT OUTER JOIN company_users AS cu 
                        ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
                WHERE
                    cl.url = 'api/company/auth/login' 
                    AND cu.email NOT LIKE '%@brownreverse%' 
                    AND DATE (cl.created_at) >= ('{current_month}' - INTERVAL 1 MONTH) 
                    AND DATE (cl.created_at) < '{current_month}' 
                GROUP BY
                    cu.id
                    , MONTH (cl.created_at)
            ) 
            SELECT
                company_id
                , sum(login_count) AS login_count 
            FROM
                month_login_temp
            GROUP BY 
                company_id
        ) 
        , year_month_previous_login AS ( 
            WITH month_login_temp AS ( 
                SELECT
                    cu.company_id AS company_id
                    , yearweek(cl.created_at) AS created_yearweek
                    , count(DISTINCT MONTH (cl.created_at)) AS login_count 
                FROM
                    company_logs AS cl 
                    LEFT OUTER JOIN company_users AS cu 
                        ON JSON_UNQUOTE(json_extract(cl.request_parameters, '$.email')) = cu.email 
                WHERE
                    cl.url = 'api/company/auth/login' 
                    AND cu.email NOT LIKE '%@brownreverse%' 
                    AND DATE (cl.created_at) >= ('{previous_month}' - INTERVAL 1 MONTH) 
                    AND DATE (cl.created_at) < '{previous_month}' 
                GROUP BY
                    cu.id
                    , MONTH (cl.created_at)
            ) 
            SELECT
                company_id
                , sum(login_count) AS login_count 
            FROM
                month_login_temp
            GROUP BY company_id
        ) 
        SELECT
            ymc.company_id
            , ymc.login_count / ymp.login_count AS monthly_login_previous_month_ratio
        FROM
            year_month_current_login AS ymc 
            LEFT OUTER JOIN year_month_previous_login AS ymp 
                ON ymc.company_id = ymp.company_id
    ) 
    , company_area_info AS ( 
        SELECT
            com.id AS company_id
            , com.name AS company_name
            , pa.id AS area_id
            , pa.name AS area_name 
        FROM
            companies AS com 
            LEFT OUTER JOIN plants AS pla 
                ON com.id = pla.company_id 
            LEFT OUTER JOIN plant_areas AS pa 
                ON pla.id = pa.plant_id
    ) 
    , marker_amount AS ( 
        SELECT
            cai.company_id
            , count(*) AS marker_amount
        FROM
            markers AS mk 
            LEFT OUTER JOIN company_area_info AS cai 
                ON mk.plant_area_id = cai.area_id 
            LEFT OUTER JOIN company_users as cu
                ON mk.company_user_id = cu.id
        WHERE
             mk.created_at < '{current_month}' and cai.company_id is not null and (cu.email not like '%brownreverse%' or cu.email is null)
        GROUP by cai.company_id
    ) 
    , new_marker_amount AS ( 
        SELECT
            cai.company_id
            , count(*) AS new_marker_amount
        FROM
            markers AS mk 
            LEFT OUTER JOIN company_area_info AS cai 
                ON mk.plant_area_id = cai.area_id 
            LEFT OUTER JOIN company_users as cu
                ON mk.company_user_id = cu.id
        WHERE
            mk.created_at < '{current_month}' 
            AND mk.created_at >= DATE_SUB('{current_month}', INTERVAL 1 MONTH) and cai.company_id is not null and (cu.email not like '%brownreverse%' or cu.email is null)
        GROUP BY
            cai.company_id
    ) 
    , machine_amount AS ( 
        SELECT
            cai.company_id
            , count(*) AS machine_amount
        FROM
            assets AS ast 
            LEFT OUTER JOIN company_area_info AS cai 
                ON ast.plant_area_id = cai.area_id
            LEFT OUTER JOIN company_users as cu
                ON ast.company_user_id = cu.id
        WHERE
            ast.created_at < '{current_month}' 
            AND cai.company_id IS NOT NULL AND (cu.email not like '%brownreverse%' or cu.email is null)
        GROUP BY
            cai.company_id
    ) 
    , new_machine_amount AS ( 
        SELECT
            cai.company_id
            , count(*) AS new_machine_amount
        FROM
            assets AS ast 
            LEFT OUTER JOIN company_area_info AS cai 
                ON ast.plant_area_id = cai.area_id 
            LEFT OUTER JOIN company_users as cu
                ON ast.company_user_id = cu.id
        WHERE
            ast.created_at < '{current_month}' 
            AND ast.created_at >= DATE_SUB('{current_month}', INTERVAL 1 MONTH) 
            AND cai.company_id IS NOT NULL 
            AND (cu.email not like '%brownreverse%' or cu.email is null)
        GROUP BY
            cai.company_id
    ) 
    , measurement_amount AS ( 
        SELECT
            cai.company_id
            , count(*) AS measurement_amount
        FROM
            measure_lengths AS ms 
            LEFT OUTER JOIN company_area_info AS cai 
                ON ms.plant_area_id = cai.area_id
            LEFT OUTER JOIN company_users as cu
                ON ms.company_user_id = cu.id
        WHERE
            ms.created_at < '{current_month}' 
            AND cai.company_id IS NOT NULL 
            AND (cu.email not like '%brownreverse%' or cu.email is null)
        GROUP BY
            cai.company_id
    ) 
    , new_measurement_amount AS ( 
        SELECT
            cai.company_id
            , count(*) AS new_measurement_amount
        FROM
            measure_lengths AS ms 
            LEFT OUTER JOIN company_area_info AS cai 
                ON ms.plant_area_id = cai.area_id 
            LEFT OUTER JOIN company_users as cu
                ON ms.company_user_id = cu.id
        WHERE
            ms.created_at < '{current_month}' 
            AND ms.created_at >= DATE_SUB('{current_month}', INTERVAL 1 MONTH) 
            AND cai.company_id IS NOT NULL 
            AND (cu.email not like '%brownreverse%' or cu.email is null)
        GROUP BY
            cai.company_id
    ) 
    , simulation_amount AS ( 
        SELECT
            cai.company_id
            , count(*) AS simulation_amount
        FROM
            plant_area_objects AS pao 
            LEFT OUTER JOIN company_area_info AS cai 
                ON pao.plant_area_id = cai.area_id 
            LEFT OUTER JOIN company_users as cu
                ON pao.company_user_id = cu.id
        WHERE
            pao.created_at < '{current_month}' 
            AND cai.company_id IS NOT NULL 
            AND (cu.email not like '%brownreverse%' or cu.email is null)
        GROUP BY
            cai.company_id
    ) 
    , new_simulation_amount AS ( 
        SELECT
            cai.company_id
            , count(*) AS new_simulation_amount
        FROM
            plant_area_objects AS pao 
            LEFT OUTER JOIN company_area_info AS cai 
                ON pao.plant_area_id = cai.area_id 
            LEFT OUTER JOIN company_users as cu
                ON pao.company_user_id = cu.id
        WHERE
            pao.created_at < '{current_month}' 
            AND pao.created_at >= DATE_SUB('{current_month}', INTERVAL 1 MONTH) 
            AND cai.company_id IS NOT NULL 
            AND (cu.email not like '%brownreverse%' or cu.email is null)
        GROUP BY
            cai.company_id
    ) 
    , pipenavi_amount AS ( 
        SELECT
            cai.company_id
            , count(*) AS pipenavi_amount
        FROM
            pipe_groups AS pg 
            LEFT OUTER JOIN company_area_info AS cai 
                ON pg.plant_area_id = cai.area_id 
            LEFT OUTER JOIN company_users AS cu
                ON pg.company_user_id = cu.id
        WHERE
            pg.created_at < '{current_month}' 
            AND cai.company_id IS NOT NULL 
            AND (cu.email not like '%brownreverse%' or cu.email is null) 
        GROUP BY
            cai.company_id
    ) 
    , new_pipenavi_amount AS ( 
        SELECT
            cai.company_id
            , count(*) AS new_pipenavi_amount
        FROM
            pipe_groups AS pg 
            LEFT OUTER JOIN company_area_info AS cai 
                ON pg.plant_area_id = cai.area_id 
            LEFT OUTER JOIN company_users AS cu
                ON pg.company_user_id = cu.id
        WHERE
            pg.created_at < '{current_month}' 
            AND pg.created_at >= DATE_SUB('{current_month}', INTERVAL 1 MONTH) 
            AND cai.company_id IS NOT NULL 
            AND (cu.email not like '%brownreverse%' or cu.email is null) 
        GROUP BY
            cai.company_id
    ) 
    SELECT 
        concat(ur.company_id,'_',ur.company_name) as '会社'
        , ur.user_amount
        , nur.user_new 
        , dla.daily_login_amount
        , concat(round(dlp.daily_login_count_ratio*100,0),'%')
        , wla.weekly_login_amount
        , concat(round(wlp.weekly_login_count_ratio*100,0),'%')
        , mla.monthly_login_amount 
        , concat(round(mlp.monthly_login_previous_month_ratio*100,0),'%')
        , ma.marker_amount
       , nma.new_marker_amount
       , mca.machine_amount
       , nmca.new_machine_amount
       , msa.measurement_amount
       , nmsa.new_measurement_amount
       , sma.simulation_amount
       , nsma.new_simulation_amount
       , pa.pipenavi_amount
       , npa.new_pipenavi_amount
       , ""
       , ""
       , ""
       , ""
       , ""
    FROM
        user_registration AS ur 
        LEFT OUTER JOIN new_user_registration AS nur 
            ON ur.company_id = nur.company_id
        LEFT OUTER JOIN daily_login_amount AS dla
            ON ur.company_id = dla.company_id
        LEFT OUTER JOIN daily_login_previous_month AS dlp
            ON ur.company_id = dlp.company_id
        LEFT OUTER JOIN weekly_login_amount AS wla
            ON ur.company_id = wla.company_id
        LEFT OUTER JOIN weekly_login_previous_month AS wlp
            ON ur.company_id = wlp.company_id
        LEFT OUTER JOIN monthly_login_amount AS mla
            ON ur.company_id = mla.company_id
        LEFT OUTER JOIN monthly_login_previous_month AS mlp
            ON ur.company_id = mlp.company_id
        LEFT OUTER JOIN marker_amount AS ma
            ON ur.company_id = ma.company_id
        LEFT OUTER JOIN new_marker_amount AS nma
            ON ur.company_id = nma.company_id
       LEFT OUTER JOIN machine_amount AS mca
           ON ur.company_id = mca.company_id
       LEFT OUTER JOIN new_machine_amount AS nmca
           ON ur.company_id = nmca.company_id
       LEFT OUTER JOIN measurement_amount AS msa
           ON ur.company_id = msa.company_id
       LEFT OUTER JOIN new_measurement_amount AS nmsa
           ON ur.company_id = nmsa.company_id
       LEFT OUTER JOIN simulation_amount AS sma
           ON ur.company_id = sma.company_id
       LEFT OUTER JOIN new_simulation_amount AS nsma
           ON ur.company_id = nsma.company_id
       LEFT OUTER JOIN pipenavi_amount AS pa
           ON ur.company_id = pa.company_id
       LEFT OUTER JOIN new_pipenavi_amount AS npa
           ON ur.company_id = npa.company_id
    '''

    company_amount = read_query(conn, q_company_amount)[0][0]

    result_df = pd.read_sql_query(q_ivr_customer, conn).fillna(0)
    result_df.set_index('会社', inplace=True)

    result_df = result_df.stack().reset_index(level=1, drop=True).to_frame(f'{year_month}')
    d = ["ユーザー登録数（合計）", "新規登録ユーザー数", "日間平均ログイン人数", "日間平均ログイン人数（前月比）",
         "週間平均ログイン"
            , "週間平均ログイン（前月比）", "月間ログイン人数", "月間ログイン人数（前月比）", "マーカー数（合計）",
         "新規マーカー数", "機番登録数（合計）", "新規機番登録数", "測長数（合計）", "新規測長数",
         "空間シミュレーション数（合計）", "新規空間シミュレーション数", "配管登録数（合計）", "新規配管登録数",
         "アプリDL数",
         "全体面積", "撮影面積", "撮影面積割合", "滞在時間"] * int(len(result_df) / 23)
    result_df.insert(0, "項目", d, allow_duplicates=False)

    return result_df

#result_df = get_analysis_result(previous_month_days,current_month_days,previous_month_weeks,current_month_weeks,previous_month,current_month)
result_df_set = {} #
for i in range(12):
    result_df_set[i] = get_analysis_result(
        calendar.monthrange(year_month_pairs[i+2][0], year_month_pairs[i+2][1])[1], # previouse_month_days
        calendar.monthrange(year_month_pairs[i+1][0], year_month_pairs[i+1][1])[1], # current_month_days
        len(calendar.monthcalendar(year_month_pairs[i+2][0], year_month_pairs[i+2][1])), # previous_month_weeks
        len(calendar.monthcalendar(year_month_pairs[i+1][0], year_month_pairs[i+1][1])), # current_month_weeks
        str(year_month_pairs[i+1][0]) + "/" + str(year_month_pairs[i+1][1]) + "/01", # previous_month
        str(year_month_pairs[i][0]) + "/" + str(year_month_pairs[i][1]) + "/01", # current_month
        str(year_month_pairs[i+1][0]) + "年" + str(year_month_pairs[i+1][1]) + "月" # year_month : column name
    )

# output result xlsx file.
# set the output file name
output_filename = str(datetime.date.today()) +  '_IVRCustomer.xlsx'
#result_df_set[0].to_excel(f'{output_filename}')
#result_df_set[1].to_excel('20240831_test.xlsx')

# for k in range(12):
#     result_df_set[k].to_csv(f'output_file_20240902_{k}')

# Merge data
output_df = result_df_set[0]
for i in range(1, 12):
    output_df = pd.merge(output_df, result_df_set[i], on=["会社", "項目"], how="left")


# Change column sequence.
column_names = list(output_df.columns.values)
column_names.reverse()
column_names.insert(0,column_names[-1])
column_names.pop()
output_df = output_df[column_names]


output_df.to_excel('IVR_customer_merged_file.xlsx')



# workbook = openpyxl.load_workbook(f'./{output_filename}')
# worksheet = workbook.active
# #
# # # font
# # font = Font(name='Calibri', size=11, bold=False)
# #
# # # Fill header with LightStellBlue
# # for cell in worksheet[1]:
# #     cell.fill = openpyxl.styles.PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid')
# #     cell.font = font
# #
# #
# # for row in worksheet.iter_rows(values_only=False):
# #     for cell in row:
# #         cell.font = font
# #
# # item_column = worksheet['B']
# # for cell in item_column:
# #     if cell.row != 1 and '前月比' in str(cell.value) and '%' in str(
# #             worksheet.cell(row=cell.row, column=cell.column + 1).value):
# #         worksheet.cell(row=cell.row, column=cell.column + 1).font = Font(color="FF0000")
#
# workbook.save(f'{output_filename}')
# workbook.close()

